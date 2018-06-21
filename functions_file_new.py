#!python3
''' tak funkcja tworzy plik excelowy'''
def create_file():

    from openpyxl import Workbook
    from openpyxl.cell import Cell
    from openpyxl.styles import Color, Fill, Font, Alignment
    import sqlite3

    conn = sqlite3.connect('budget.sqlite')  #łączy się z naszym plikiem
    cur = conn.cursor()

    global raport
    raport = Workbook() 

    cur.execute('SELECT * FROM Employees')
    table1 = cur.fetchall()
    for row in table1:
        ws = raport.create_sheet(row[0])

        cur.execute('SELECT * FROM Headers') 
        table2 = cur.fetchall()
        for row in table2:
            #print(row)
            ws.cell(column=int(row[0]), row=1).value = row[1]

    ft = Font(name='Arial',size=8, italic=True)
    wt = Alignment(wrap_text=True, horizontal='center', vertical='center')
    for col in range (1,19):
        ws.cell(column=col, row=1).font = ft
        ws.cell(column=col, row=1).alignment = wt

    ws.freeze_panes = 'A2'

    global total 
    total = raport['total']
    std=raport.get_sheet_by_name('Sheet')
    raport.remove_sheet(std)
    raport.save('budget_tk_et.xlsx')
    
    return raport
    return total

'''ta funkcja otwiera plik excelowy'''
def open_file():
    import os
    plik_xlsx = r"C:\Users\Monika\Desktop\eoffice\budget\budget_tk_et.xlsx\\" 
    os.startfile(plik_xlsx)


''' ta funkcja pobiera dane z xmla'''
def eoffice_et():

    import xml.etree.ElementTree as ET
    from openpyxl import Workbook
    from openpyxl.cell import Cell
    from openpyxl.styles import Color, Fill
    from openpyxl.styles import Font, Alignment

    from datetime import date
    import os

    counter_wierszy = 2

    pierwszy_plik = pierwszy_tekst.get()
    print('pierwszy plik: ', pierwszy_plik) 
    if not os.path.exists("C:/Users/Monika/Desktop/eoffice/budget/testi24 " + pierwszy_plik + ".xml"):
        print("Plik o nazwie testi24 " + pierwszy_plik + ".xml nie istnieje. Spróbuj ponownie.")
        #return pierwszy_plik
        quit()

    ostatni_plik = ostatni_tekst.get()
    print('ostatni plik: ', ostatni_plik)
    if not os.path.exists("C:/Users/Monika/Desktop/eoffice/budget/testi24 " + ostatni_plik + ".xml"):
        print("Plik o nazwie testi24 " + ostatni_plik + ".xml nie istnieje. Spróbuj ponownie.")
        #return ostatni_plik
        quit()

    zakres = range(int(pierwszy_plik), int(ostatni_plik)+1)

    for i in zakres:
        filename = r"C:\Users\Monika\Desktop\eoffice\\"
        numer_pliku = str(i)
        nazwa_pliku = "testi24 " + numer_pliku + ".xml"
        filename += nazwa_pliku
        tree = ET.parse(filename)
        root = tree.getroot()
        
        for invoice in root.findall('purchaseInvoice'):
            
            PZ = []
            for row in invoice:            
                if row.tag =='handlingEvents':
                    for item in row:
                        if item.tag == 'handlingEvent':
                            for event in item:
                                if event.tag == 'handlingComment':
                                    opis_eventu = str(event.text)
                                    if str(opis_eventu)[:14] == "Attachment 'PZ":
                                        spacja_po_PZ = str(opis_eventu)[14]
                                        if spacja_po_PZ == " ":
                                            początek = 15
                                        else:
                                            początek = 14
                                        try:
                                            koniec = opis_eventu.find('2018') + 4
                                        except:
                                            koniec = 26
                                        data_PZ = opis_eventu[początek:koniec]
                                        PZ.append(data_PZ)
                                        #print(PZ)
                    #slownik_Review = []
                    for item in row:
                        if item.tag == 'handlingEvent':
                            for event in item:
                                if event.tag == 'handlingPhaseName':
                                    if event.text == "Review":
                                        for event in item:

                                            if event.tag == 'handlerName':
                                                review_person = event.text
                                                print(review_person)
            for row in invoice:
                if row.tag == 'invoiceNumber':
                    numer_faktury = row.text
            for row in invoice:
                if row.tag == 'invoiceDate':
                    data_faktury = row.text
            for row in invoice:
                if row.tag == 'supplier':
                    for item in row:
                        if item.tag == "supplierNumber":
                            numer_dostawcy = item.text
            for row in invoice:
                if row.tag == 'supplier':
                    for item in row:
                        if item.tag == "supplierName":
                            nazwa_dostawcy = item.text
            for row in invoice:            
                if row.tag == "invoiceCurrency":
                    waluta = row.text
            for row in invoice:
                if row.tag == "voucherDateCurrencyRate":
                    kurs_waluty_raw = "%5.4f" % float(row.text)
                    kurs_waluty = kurs_waluty_raw.replace(".", ",")


            for row in invoice:
                if row.tag == "invoiceMessage":
                    tabela = row.text
                    #print(tabela)

            

            arkusz = raport[review_person]
            for row in invoice:
                lista_opisów = []
                lista_opisów.append(numer_faktury)
                if row.tag == 'invoicePosting':
                    for entry in row:
                        if entry.tag == "accountingEntry":
                            
                            for item in entry:
                                if item.tag == "accountingEntryRow":
                                    for entry in item:
                                        if entry.tag == "entryRowDescription":
                                            opis = entry.text
                                            cost_center = opis[-1]
                                            #print("cc", cost_center)
                                        if entry.tag == "entryRowAmountType":
                                            if entry.text == "0":
                                                mnożnik = 1
                                            else:
                                                mnożnik = -1
                                        if entry.tag == "entryRowNetAmount":
                                            kwota = float(entry.text) * mnożnik
                                        if entry.tag == "ledgerAccountNumber":
                                            if entry.text[:2] == "40" or entry.text[:3] == "412":
                                                lista_opisów.append(sl_pracownicy[review_person])
                                                lista_opisów.append(opis)
                                                #print(lista_opisów)
                                                
                                                '''#wrzucanie do arkuszy pracowników'''
                                                
                                                arkusz.cell(column=1, row=sl_pracownicy[review_person]).value = numer_pliku
                                                arkusz.cell(column=2, row=sl_pracownicy[review_person]).value = numer_faktury
                                                arkusz.cell(column=3, row=sl_pracownicy[review_person]).value = data_faktury
                                                arkusz.cell(column=4, row=sl_pracownicy[review_person]).value = numer_dostawcy
                                                arkusz.cell(column=6, row=sl_pracownicy[review_person]).value = nazwa_dostawcy
                                                if nazwa_dostawcy == "Reka Kumi OY":
                                                    arkusz.cell(column=5, row=sl_pracownicy[review_person]).value = 1
                                                else:
                                                    arkusz.cell(column=5, row=sl_pracownicy[review_person]).value = 2
                                                arkusz.cell(column=7, row=sl_pracownicy[review_person]).value = opis
                                                arkusz.cell(column=8, row=sl_pracownicy[review_person]).value = waluta
                                                arkusz.cell(column=9, row=sl_pracownicy[review_person]).value = kwota
                                                arkusz.cell(column=11, row=sl_pracownicy[review_person]).value = kurs_waluty
                                                arkusz.cell(column=12, row=sl_pracownicy[review_person]).value = tabela
                                                arkusz.cell(column=13, row=sl_pracownicy[review_person]).value = ("%7.2f" % (float(kwota) * float(kurs_waluty_raw))).replace('.', ',')
                                                try:
                                                    arkusz.cell(column=14, row=sl_pracownicy[review_person]).value = PZ[0]
                                                except:
                                                    arkusz.cell(column=14, row=sl_pracownicy[review_person]).value = 'puste'
                                                arkusz.cell(column=17, row=sl_pracownicy[review_person]).value = entry.text
                                                arkusz.cell(column=18, row=sl_pracownicy[review_person]).value = review_person

                                                '''wrzucanie do arkusza "razem" '''

                                                total.cell(column=1, row=counter_wierszy).value = numer_pliku
                                                total.cell(column=2, row=counter_wierszy).value = numer_faktury
                                                total.cell(column=3, row=counter_wierszy).value = data_faktury
                                                total.cell(column=4, row=counter_wierszy).value = numer_dostawcy
                                                total.cell(column=6, row=counter_wierszy).value = nazwa_dostawcy
                                                if nazwa_dostawcy == "Reka Kumi OY":
                                                    total.cell(column=5, row=counter_wierszy).value = 1
                                                else:
                                                    total.cell(column=5, row=counter_wierszy).value = 2
                                                total.cell(column=7, row=counter_wierszy).value = opis
                                                total.cell(column=8, row=counter_wierszy).value = waluta
                                                total.cell(column=9, row=counter_wierszy).value = kwota
                                                total.cell(column=11, row=counter_wierszy).value = kurs_waluty
                                                total.cell(column=12, row=counter_wierszy).value = tabela
                                                total.cell(column=13, row=counter_wierszy).value = ("%7.2f" % (float(kwota) * float(kurs_waluty_raw))).replace('.', ',')
                                                try:
                                                    total.cell(column=14, row=counter_wierszy).value = PZ[0]
                                                except:
                                                    total.cell(column=14, row=counter_wierszy).value = 'puste'
                                                # arkusz.cell(column=15, row=sl_pracownicy[review_person]).value = cost_center
                                                # slownik_opisów = {"miesz":4203, "guma ":4203, "silik":4203, "nyro ":4203, "eleme":4205, 
                                                #                 "uchwyt":4205, "obejm":4205, "etyki":4251, "opako":4251, "folia":4251,
                                                #                 "palet":4251, "karto":4251, "nici ":4291, "rekaw":4291, "klej ":4291, 
                                                #                 "nakle":4291, "zawie":6504, "tasma":4291}
                                                # try: 
                                                #     arkusz.cell(column=16, row=sl_pracownicy[review_person]).value = slownik_opisów[opis[:5]]
                                                # except:
                                                #     arkusz.cell(column=16, row=sl_pracownicy[review_person]).value = opis
                                                total.cell(column=17, row=counter_wierszy).value = entry.text
                                                total.cell(column=18, row=counter_wierszy).value = review_person
                                                sl_pracownicy[review_person] +=1
                                                counter_wierszy += 1
                                                print(lista_opisów)
        total.cell(column=21,row=1).value = "od pliku "+str(pierwszy_plik)
        total.cell(column=22,row=1).value = "do pliku "+str(ostatni_plik)

        raport.save('eoffice_et.xlsx')
    print("zakończono")

    # sl_pracownicy = {'total':2, 'Durska, Agata':2,'Perska, Anna':2,'Okrajek, Krzysztof':2,'Mackowska, Magdalena ':2,
    #                'Jankowiak, Monika':2,'Zwierzynska, Angelika':2,'Bobrow, Jakub':2,'Peters, Piotr':2,
    #                'Kieronska, Maria':2,'Golinski, Radoslaw':2,'Brach, Jakub':2, 'Staroscinska, Maja':2, 'Walich, Piotr':2}
    # sl_naglowki = {'1':'Numer pliku testi','2':'Numer faktury','3':'Data faktury','4':'ID dostawcy','5':'1=IC   2=3rd Party',
    #                 '6':'Nazwa dostawcy','7':'Przedmiot dostawy','8':'waluta','9':'kwota w walucie','10':'data wyceny','11':'kurs księgi',
    #                 '12':'tabela','13':'kwota w zł księgi','14':'data wpływu na magazyn','15':'centrum kosztowe','16':'konto Epicor',
    #                 '17':'konto Symfonia','18':'Review'}

    # for pracownik in sl_pracownicy:
    #     ws = raport.create_sheet(pracownik)
    #     for naglowek in sl_naglowki:
    #         ws.cell(column=int(naglowek), row=1).value = sl_naglowki[naglowek]