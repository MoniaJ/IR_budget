#!python3
"""
Ten plik pyta w GUI o zakres plików testi do przerobienia - pierwszy numer pliku i ostatni.
"""
#How to Open and Close Programs: https://www.youtube.com/watch?v=IJm9m8kv7gU
def eoffice_et():

    import xml.etree.ElementTree as ET
    from openpyxl import Workbook
    from openpyxl.cell import Cell
    from openpyxl.styles import Color, Fill
    from openpyxl.styles import Font, Alignment

    from datetime import date
    import os

    #przygotowanie pliku excel:
    raport = Workbook() 
    arkusz = raport.active
    arkusz.title = 'arkusz 1'
    

    arkusz.cell(column=1, row=1).value = 'Numer pliku testi'
    arkusz.cell(column=2, row=1).value = 'Numer faktury'
    arkusz.cell(column=3, row=1).value = 'Data faktury'
    arkusz.cell(column=4, row=1).value = 'ID dostawcy'
    arkusz.cell(column=5, row=1).value = '1=IC   2=3rd Party'
    arkusz.cell(column=6, row=1).value = 'Nazwa dostawcy'
    arkusz.cell(column=7, row=1).value = 'Przedmiot dostawy'
    arkusz.cell(column=8, row=1).value = 'waluta'
    arkusz.cell(column=9, row=1).value = 'kwota w walucie'
    arkusz.cell(column=10, row=1).value = 'data wyceny'
    arkusz.cell(column=11, row=1).value = 'kurs księgi'
    arkusz.cell(column=12, row=1).value = 'tabela'
    arkusz.cell(column=13, row=1).value = 'kwota w zł księgi'
    arkusz.cell(column=14, row=1).value = 'data wpływu na magazyn'
    arkusz.cell(column=15, row=1).value = 'centrum kosztowe'
    arkusz.cell(column=16, row=1).value = 'konto'

    ft = Font(name='Arial',size=8, italic=True)
    wt = Alignment(wrap_text=True, horizontal='center', vertical='center')
    for col in range (1,17):
        arkusz.cell(column=col, row=1).font = ft
        arkusz.cell(column=col, row=1).alignment = wt

    arkusz.freeze_panes = 'A2'

    counter_wierszy = 2

    pierwszy_plik = pierwszy_tekst.get()
    print('pierwszy plik: ', pierwszy_plik) 
    if not os.path.exists("C:/Users/Monika/Desktop/eoffice/testi24 " + pierwszy_plik + ".xml"):
        print("Plik o nazwie testi24 " + pierwszy_plik + ".xml nie istnieje. Spróbuj ponownie.")
        quit()
    ostatni_plik = ostatni_tekst.get()
    print('ostatni plik: ', ostatni_plik)
    if not os.path.exists("C:/Users/Monika/Desktop/eoffice/testi24 " + ostatni_plik + ".xml"):
        print("Plik o nazwie testi24 " + ostatni_plik + ".xml nie istnieje. Spróbuj ponownie.")
        quit()

    zakres = range(int(pierwszy_plik), int(ostatni_plik)+1)

    for i in zakres:
        filename = r"C:\Users\Monika\Desktop\eoffice\\"
        numer_pliku = str(i)
        nazwa_pliku = "testi24 " + numer_pliku + ".xml"
        filename += nazwa_pliku
        tree = ET.parse(filename)
        root = tree.getroot()
        # counter_faktur = 0
        # for invoice in root.findall('purchaseInvoice'):
        #     counter_faktur += 1
        # print('ilość faktur: ', counter_faktur)
        # arkusz.cell(column=18, row=1).value = 'Ilość faktur: ' + str(counter_faktur)

        for invoice in root.findall('purchaseInvoice'):
            
            for row in invoice:
                if row.tag == 'invoiceNumber':
                    numer_faktury = row.text
            for row in invoice:
                if row.tag == 'invoiceDate':
                    data_faktury = row.text
            for row in invoice:
                if row.tag == 'supplier':
                    for item in row:
                        if item.tag == "supplierNumber":
                            numer_dostawcy = item.text
            for row in invoice:
                if row.tag == 'supplier':
                    for item in row:
                        if item.tag == "supplierName":
                            nazwa_dostawcy = item.text
            for row in invoice:            
                if row.tag == "invoiceCurrency":
                    waluta = row.text
            for row in invoice:
                if row.tag == "voucherDateCurrencyRate":
                    kurs_waluty_raw = "%5.4f" % float(row.text)
                    kurs_waluty = kurs_waluty_raw.replace(".", ",")


            for row in invoice:
                if row.tag == "invoiceMessage":
                    tabela = row.text
                    #print(tabela)

            PZ = []
            for row in invoice:            
                if row.tag =='handlingEvents':
                    for item in row:
                        if item.tag == 'handlingEvent':
                            for event in item:
                                if event.tag == 'handlingComment':
                                    opis_eventu = str(event.text)
                                    if str(opis_eventu)[:14] == "Attachment 'PZ":
                                        spacja_po_PZ = str(opis_eventu)[14]
                                        if spacja_po_PZ == " ":
                                            początek = 15
                                        else:
                                            początek = 14
                                        try:
                                            koniec = opis_eventu.find('2018') + 4
                                        except:
                                            koniec = 26
                                        data_PZ = opis_eventu[początek:koniec]
                                        PZ.append(data_PZ)
                                        #print(PZ)
            for row in invoice:
                lista_opisów = []
                lista_opisów.append(numer_faktury)
                if row.tag == 'invoicePosting':
                    for entry in row:
                        if entry.tag == "accountingEntry":
                            
                            for item in entry:
                                if item.tag == "accountingEntryRow":
                                    for entry in item:
                                        if entry.tag == "entryRowDescription":
                                            opis = entry.text
                                            cost_center = opis[-1]
                                            #print("cc", cost_center)
                                        if entry.tag == "entryRowAmountType":
                                            if entry.text == "0":
                                                mnożnik = 1
                                            else:
                                                mnożnik = -1
                                        if entry.tag == "entryRowNetAmount":
                                            kwota = float(entry.text) * mnożnik
                                        if entry.tag == "ledgerAccountNumber":
                                            if entry.text == "301" or entry.text == "305":                                            
                                                lista_opisów.append(counter_wierszy)
                                                lista_opisów.append(opis)
                                                #print(lista_opisów)
                                                arkusz.cell(column=1, row=counter_wierszy).value = numer_pliku
                                                arkusz.cell(column=2, row=counter_wierszy).value = numer_faktury
                                                arkusz.cell(column=3, row=counter_wierszy).value = data_faktury
                                                arkusz.cell(column=4, row=counter_wierszy).value = numer_dostawcy
                                                arkusz.cell(column=6, row=counter_wierszy).value = nazwa_dostawcy
                                                if nazwa_dostawcy == "Reka Kumi OY":
                                                    arkusz.cell(column=5, row=counter_wierszy).value = 1
                                                else:
                                                    arkusz.cell(column=5, row=counter_wierszy).value = 2
                                                arkusz.cell(column=7, row=counter_wierszy).value = opis
                                                arkusz.cell(column=8, row=counter_wierszy).value = waluta
                                                arkusz.cell(column=9, row=counter_wierszy).value = kwota
                                                arkusz.cell(column=11, row=counter_wierszy).value = kurs_waluty
                                                arkusz.cell(column=12, row=counter_wierszy).value = tabela
                                                arkusz.cell(column=13, row=counter_wierszy).value = ("%7.2f" % (float(kwota) * float(kurs_waluty_raw))).replace('.', ',')
                                                try:
                                                    arkusz.cell(column=14, row=counter_wierszy).value = PZ[0]
                                                except:
                                                    arkusz.cell(column=14, row=counter_wierszy).value = 'puste'
                                                arkusz.cell(column=15, row=counter_wierszy).value = cost_center
                                                slownik_opisów = {"miesz":4203, "guma ":4203, "silik":4203, "nyro ":4203, "eleme":4205, 
                                                                "uchwyt":4205, "obejm":4205, "etyki":4251, "opako":4251, "folia":4251,
                                                                "palet":4251, "karto":4251, "nici ":4291, "rekaw":4291, "klej ":4291, 
                                                                "nakle":4291, "zawie":6504, "tasma":4291}
                                                try: 
                                                    arkusz.cell(column=16, row=counter_wierszy).value = slownik_opisów[opis[:5]]
                                                except:
                                                    arkusz.cell(column=16, row=counter_wierszy).value = opis
                                                counter_wierszy +=1
            
        raport.save('eoffice_et.xlsx')
    print("zakończono")


#https://www.tutorialspoint.com/python/python_gui_programming.htm
#https://docs.python.org/3.6/library/tkinter.html
#https://www.python-course.eu/tkinter_entry_widgets.php

import tkinter
from tkinter import *

root = Tk()
root.minsize(700,60)
root.title("Pobieranie danych z IR do pliku 301")
root.configure(background = "blue")
frame = Frame(root)

frame.pack()
pierwszy = Label(frame, text="Wprowadź trzycyfrowy numer pierwszego pliku testi:", font = "Helvetica 13 italic", bg = "blue", fg = "white")
pierwszy_tekst = Entry(frame, bd = 3)
pierwszy.pack(side = LEFT)

pierwszy_tekst.pack(side = RIGHT)
pierwszy_tekst.focus_set()
#focus_set umieszcza kursor w pierwszym okienku

midframe = Frame(root)
midframe.pack(side = TOP)
ostatni = Label(midframe, text="Wprowadź trzycyfrowy numer ostatniego pliku testi:", font = "Helvetica 13 italic", bg = "blue", fg = "white")
ostatni_tekst = Entry(midframe, bd = 3)
ostatni.pack(side = LEFT)
ostatni_tekst.pack(side = RIGHT)

bottomframe = Frame(root)
#bottomframe.pack(side = BOTTOM)
bottomframe.pack()
B = tkinter.Button(bottomframe, text ="Uruchom", font = "Helvetica 13 italic", command = eoffice_et, bd = 4)
#B.pack(side = BOTTOM)
B.pack()

exitButton = Button(root, text = "Exit", command = root.destroy).pack(side = BOTTOM)
#bd = border
mainloop()

#https://www.youtube.com/watch?v=tMhFk_GgmFk
